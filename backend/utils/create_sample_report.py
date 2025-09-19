"""
创建示例报表文件
用于在没有实际报表时提供默认展示
"""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_sample_report():
    """创建一个示例报表Excel文件"""
    
    # 输出路径
    output_dir = Path(__file__).parent.parent / "reports" / "api_generated"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 固定的UUID作为默认报表ID
    sample_report_id = "00000000-0000-0000-0000-000000000001"
    output_file = output_dir / f"report_{sample_report_id}.xlsx"
    
    # 创建工作簿
    wb = Workbook()
    
    # ==================== 创建稳定状态参数汇总表 ====================
    ws_stable = wb.active
    ws_stable.title = "稳定状态参数汇总"
    
    # 标题
    ws_stable['A1'] = "示例报表 - 稳定状态参数汇总"
    ws_stable['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
    ws_stable['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_stable['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_stable.merge_cells('A1:G1')
    ws_stable.row_dimensions[1].height = 30
    
    # 说明文字
    ws_stable['A2'] = "说明：这是一个示例报表，用于展示系统功能。请通过AI对话上传真实数据并生成您自己的报表。"
    ws_stable['A2'].font = Font(name='微软雅黑', size=10, italic=True, color='666666')
    ws_stable.merge_cells('A2:G2')
    ws_stable.row_dimensions[2].height = 25
    
    # 表头
    headers = ['通道名称', '平均值', '标准差', '最小值', '最大值', '样本数', '单位']
    for col, header in enumerate(headers, start=1):
        cell = ws_stable.cell(row=4, column=col, value=header)
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 示例数据
    sample_data = [
        ['温度传感器-01', 25.3, 0.15, 25.0, 25.6, 1000, '℃'],
        ['压力传感器-01', 101.2, 0.8, 99.5, 102.8, 1000, 'kPa'],
        ['流量计-01', 15.7, 0.3, 15.1, 16.3, 1000, 'L/min'],
        ['液位计-01', 78.5, 1.2, 76.0, 81.0, 1000, 'mm'],
        ['电压传感器-01', 220.5, 2.5, 215.0, 225.0, 1000, 'V'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_stable.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 数值格式化
            if isinstance(value, (int, float)) and col_idx > 1 and col_idx < 7:
                cell.number_format = '0.00' if col_idx != 6 else '0'
    
    # 设置列宽
    ws_stable.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws_stable.column_dimensions[col].width = 12
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    for row in ws_stable.iter_rows(min_row=4, max_row=4+len(sample_data), 
                                    min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # ==================== 创建功能计算结果表 ====================
    ws_calc = wb.create_sheet("功能计算结果")
    
    # 标题
    ws_calc['A1'] = "示例报表 - 功能计算结果"
    ws_calc['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
    ws_calc['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    ws_calc['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_calc.merge_cells('A1:E1')
    ws_calc.row_dimensions[1].height = 30
    
    # 说明
    ws_calc['A2'] = "以下是基于示例数据计算的功能指标"
    ws_calc['A2'].font = Font(name='微软雅黑', size=10, italic=True, color='666666')
    ws_calc.merge_cells('A2:E2')
    
    # 表头
    calc_headers = ['指标名称', '计算结果', '单位', '计算公式', '备注']
    for col, header in enumerate(calc_headers, start=1):
        cell = ws_calc.cell(row=4, column=col, value=header)
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 示例计算数据
    calc_data = [
        ['能效比', 3.45, '-', 'COP = Q_out / W_in', '制冷性能指标'],
        ['系统效率', 87.3, '%', 'η = (P_out / P_in) × 100', '能量转换效率'],
        ['流量积分', 942.5, 'L', '∫ Q(t) dt', '累计流量'],
        ['温度波动', 0.59, '℃', 'ΔT = T_max - T_min', '温度稳定性'],
        ['平均功率', 1.25, 'kW', 'P_avg = ΣP / n', '平均能耗'],
    ]
    
    for row_idx, row_data in enumerate(calc_data, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_calc.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 数值格式化
            if isinstance(value, (int, float)):
                cell.number_format = '0.00'
            
            cell.border = thin_border
    
    # 设置列宽
    ws_calc.column_dimensions['A'].width = 18
    ws_calc.column_dimensions['B'].width = 12
    ws_calc.column_dimensions['C'].width = 10
    ws_calc.column_dimensions['D'].width = 25
    ws_calc.column_dimensions['E'].width = 20
    
    # ==================== 创建报表信息页 ====================
    ws_info = wb.create_sheet("报表信息")
    
    # 标题
    ws_info['A1'] = "报表元数据"
    ws_info['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
    ws_info['A1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    ws_info['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_info.merge_cells('A1:B1')
    ws_info.row_dimensions[1].height = 30
    
    # 报表信息
    info_data = [
        ['报表ID', sample_report_id],
        ['报表名称', '系统示例报表'],
        ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['报表类型', '稳定状态参数汇总 + 功能计算'],
        ['数据来源', '示例数据'],
        ['通道数量', 5],
        ['数据点数', 1000],
        ['系统版本', '1.0.0'],
        ['', ''],
        ['重要提示', '这是一个示例报表，用于演示系统功能'],
        ['', '要生成真实报表，请:'],
        ['步骤1', '访问AI对话页面'],
        ['步骤2', '上传您的数据文件(CSV或Excel)'],
        ['步骤3', '与AI对话配置报表参数'],
        ['步骤4', '生成并下载您的专属报表'],
    ]
    
    for row_idx, (key, value) in enumerate(info_data, start=3):
        cell_a = ws_info.cell(row=row_idx, column=1, value=key)
        cell_b = ws_info.cell(row=row_idx, column=2, value=value)
        
        if row_idx == 3 or (row_idx >= 10 and key):
            cell_a.font = Font(name='微软雅黑', size=10, bold=True)
        else:
            cell_a.font = Font(name='微软雅黑', size=10)
        
        cell_b.font = Font(name='微软雅黑', size=10)
        
        if row_idx >= 10:
            cell_a.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            cell_b.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 50
    
    # 保存文件
    wb.save(output_file)
    print(f"[OK] 示例报表已创建: {output_file}")
    print(f"   报表ID: {sample_report_id}")
    print(f"   文件大小: {output_file.stat().st_size} bytes")
    
    return sample_report_id, output_file

if __name__ == "__main__":
    create_sample_report()

