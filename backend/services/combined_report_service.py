"""
合并报表服务 - 生成稳态、功能计算、状态评估三张表并合并到一个Excel
"""
from pathlib import Path
from typing import Optional
import uuid
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from backend.services.steady_state_service import SteadyStateService
from backend.services.functional_service import FunctionalService
from backend.services.status_evaluation_service import StatusEvaluationService

logger = logging.getLogger(__name__)


class CombinedReportService:
    """合并报表服务"""

    def __init__(self):
        self.steady_service = SteadyStateService()
        self.functional_service = FunctionalService()
        self.status_eval_service = StatusEvaluationService()

    def generate_all_and_merge(
        self,
        steady_config_path: str,
        functional_config_path: str,
        status_eval_config_path: str,
        input_file_path: str,
        output_file_path: str
    ) -> str:
        """
        生成三张表并合并到一个Excel（单个工作表，纵向拼接 - 方案A）
        """
        output_path = Path(output_file_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # 先分别生成三个xlsx临时文件
        tmp_dir = output_path.parent / f"tmp-{uuid.uuid4()}"
        tmp_dir.mkdir(parents=True, exist_ok=True)

        steady_tmp = tmp_dir / "steady_state.xlsx"
        functional_tmp = tmp_dir / "functional.xlsx"
        status_eval_tmp = tmp_dir / "status_eval.xlsx"

        logger.info("开始生成三类报表（独立xlsx）...")
        # 稳态
        self.steady_service.generate_report(steady_config_path, input_file_path, str(steady_tmp))
        # 功能计算（获取计算结果）
        functional_results = None
        try:
            result = self.functional_service.generate_report(functional_config_path, input_file_path, str(functional_tmp))
            # 如果返回的是字典，提取calculator
            if isinstance(result, dict):
                calculator = result.get('calculator')
                if calculator and hasattr(calculator, 'results'):
                    functional_results = calculator.results
                    logger.info(f"获取到功能计算结果，共{len(functional_results)}条记录")
        except Exception as e:
            logger.warning(f"功能计算生成失败或无法获取结果: {e}")
            # 回退到简单接口
            try:
                self.functional_service.generate_report_simple(functional_config_path, input_file_path, str(functional_tmp))
            except Exception:
                pass
        # 状态评估（传递功能计算结果）
        self.status_eval_service.generate_report(status_eval_config_path, input_file_path, str(status_eval_tmp), functional_results)

        logger.info("三类报表生成完成，开始合并为单一xlsx（单sheet，纵向拼接）...")
        # 创建目标工作簿（单个sheet）
        merged_wb = Workbook()
        ws = merged_wb.active
        ws.title = "合并报表"

        # 依次将三个文件的首个工作表数据纵向拼接到同一个sheet
        current_row = 1

        # 总抬头
        title_cell = ws.cell(row=current_row, column=1, value="XX车台XX型号XX号机动态试验数据报表")
        # 12号 黑体 加粗
        title_cell.font = Font(name="SimHei", size=12, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 2  # 空一行分隔

        def append_section(title: str, src_path: Path, start_row: int) -> int:
            # 标题行
            title_cell = ws.cell(row=start_row, column=1, value=title)
            title_cell.font = Font(size=12, bold=True)
            title_cell.alignment = Alignment(horizontal='left', vertical='center')
            next_row = start_row + 1
            # 复制数据
            src_wb = load_workbook(str(src_path), data_only=True)
            src_ws: Worksheet = src_wb.worksheets[0]
            max_col = src_ws.max_column
            rows_copied = 0
            for row in src_ws.iter_rows(values_only=True):
                for col_idx, cell_val in enumerate(row, start=1):
                    ws.cell(row=next_row, column=col_idx, value=cell_val)
                next_row += 1
                rows_copied += 1
            # 第一行（表头）样式：加粗、居中、背景色
            if next_row > start_row + 1:
                header_row_index = start_row + 1
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=header_row_index, column=c)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                # 只对该表格区域加细边框（包括表头+数据行）
                thin = Side(style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                first_row = header_row_index
                last_row = header_row_index + rows_copied - 1
                for r in range(first_row, last_row + 1):
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = border
                        # 表格所有填写内容居中
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        # 仅对“表3 状态评估表”的数据行进行是/否底色标记
                        if "状态评估" in title and r > header_row_index:
                            val = cell.value
                            if isinstance(val, str):
                                text = val.strip()
                                if text == "是":
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
                                elif text == "否":
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色
            # 空一行分隔
            next_row += 1
            return next_row

        current_row = append_section("表1 稳定状态个动态参数汇总表", steady_tmp, current_row)
        current_row = append_section("表2 功能计算汇总表", functional_tmp, current_row)
        current_row = append_section("表3 状态评估表", status_eval_tmp, current_row)

        # 不对整张表设置边框，只保留各表格区域边框

        merged_wb.save(str(output_path))
        logger.info(f"合并报表已生成：{output_path}")

        # 清理临时目录（保留以便排错也可不删，这里尝试删除失败不影响）
        try:
            for p in tmp_dir.iterdir():
                p.unlink(missing_ok=True)
            tmp_dir.rmdir()
        except Exception:
            pass

        return str(output_path)

    # 方案A不再需要新增sheet复制的方法，保留占位以便未来扩展（多sheet方案B）
    def _append_first_sheet_as_new(self, src_wb, dst_wb: Workbook, new_title: Optional[str] = None):
        pass


