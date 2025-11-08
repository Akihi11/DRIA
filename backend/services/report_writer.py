"""
报表生成器 - 将快照数据写入Excel并生成图表
"""
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
from typing import List, Dict, Any, TYPE_CHECKING
import logging

if TYPE_CHECKING:
    from backend.services.status_evaluation_calculator import EvaluationItem

logger = logging.getLogger(__name__)


class ReportWriter:
    """报表生成器"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def create_report(self, snapshots: List[Dict[str, Any]], output_path: str):
        """
        创建报表文件
        
        Args:
            snapshots: 快照列表，每个快照包含timestamp和data
            output_path: 输出文件路径
        """
        # 创建Excel工作簿
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "稳定状态快照"
        
        if not snapshots:
            logger.warning("没有快照数据")
            self.workbook.save(output_path)
            return
        
        # 获取所有显示通道
        display_channels = list(snapshots[0]['data'].keys())
        
        # 写入表头
        headers = ['时间'] + display_channels
        self.worksheet.append(headers)
        
        # 格式化表头：居中、加粗、背景色
        for col_idx in range(1, len(headers) + 1):
            cell = self.worksheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 写入数据（确保所有数据都是数值格式）
        for snapshot in snapshots:
            timestamp = float(snapshot['timestamp'])  # 确保时间戳是数值
            row_values = [timestamp] + [float(snapshot['data'][ch]) for ch in display_channels]
            self.worksheet.append(row_values)
            
            # 确保Excel单元格格式为数值格式
            current_row = self.worksheet.max_row
            # 时间列设置为数值格式
            time_cell = self.worksheet.cell(row=current_row, column=1)
            time_cell.number_format = '0.000'
            # 通道数据列设置为数值格式
            for col_idx, _ in enumerate(display_channels, start=2):
                data_cell = self.worksheet.cell(row=current_row, column=col_idx)
                data_cell.number_format = '0.00'
        
        logger.info(f"已写入 {len(snapshots)} 行数据")
        
        # 生成图表 - 已关闭
        # self._create_chart(display_channels, len(snapshots))
        
        # 保存文件
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)
        logger.info(f"报表已保存: {output_path}")
    
    def _create_chart(self, channels: List[str], num_rows: int):
        """创建折线图"""
        if num_rows <= 0:
            return
        
        try:
            # 创建图表
            chart = ScatterChart()
            chart.style = 10
            
            # 设置X轴标题和格式
            chart.x_axis.title = '时间 (秒)'
            # 确保X轴显示刻度标签
            chart.x_axis.delete = False
            
            # 设置Y轴标题和格式
            chart.y_axis.title = '数值'
            # 确保Y轴显示刻度标签
            chart.y_axis.delete = False
            
            # 移除所有网格线（只保留坐标轴）
            # 在 openpyxl 中，通过直接操作 XML 元素来移除网格线
            # 设置网格线为空，使其不显示
            try:
                from openpyxl.chart.axis import Gridlines
                
                # 对于 X 轴：移除主要和次要网格线
                # 创建空的 Gridlines 对象，然后通过设置其属性来隐藏
                empty_gridlines = Gridlines()
                # 通过设置 spPr (形状属性) 来隐藏网格线
                if hasattr(empty_gridlines, 'spPr'):
                    empty_gridlines.spPr = None
                
                chart.x_axis.majorGridlines = None
                chart.x_axis.minorGridlines = None
                
                # 对于 Y 轴：移除主要和次要网格线
                chart.y_axis.majorGridlines = None
                chart.y_axis.minorGridlines = None
            except Exception as e:
                logger.warning(f"设置网格线失败: {e}")
                # 如果直接设置为 None 不起作用，尝试使用空对象
                try:
                    from openpyxl.chart.axis import Gridlines
                    empty_grid = Gridlines()
                    chart.x_axis.majorGridlines = empty_grid
                    chart.y_axis.majorGridlines = empty_grid
                except:
                    pass
            
            # 为每个通道创建一个系列
            # 注意：我们要创建非公用Y轴的图表
            # 这里简化处理，将所有系列添加到同一个图表
            for idx, channel in enumerate(channels):
                # X轴数据：时间列（第1列），跳过表头
                x_values = Reference(self.worksheet, min_col=1, min_row=2, max_row=num_rows + 1)
                
                # Y轴数据：对应通道列（注意：min_row=2跳过表头，只包含数据）
                y_values = Reference(self.worksheet, min_col=idx + 2, min_row=2, max_row=num_rows + 1)
                
                # 创建系列
                series = Series(y_values, x_values, title=channel)
                chart.series.append(series)
            
            # 设置图表位置和大小
            chart.width = 15
            chart.height = 10
            
            # 将图表添加到工作表
            chart_cell = f"A{num_rows + 3}"
            self.worksheet.add_chart(chart, chart_cell)
            
            # 在图表添加后，通过 XML 操作移除网格线（更可靠的方法）
            try:
                # 获取图表对象对应的 XML 元素
                chart_element = chart._chart
                if hasattr(chart_element, 'plotArea'):
                    plot_area = chart_element.plotArea
                    # 移除 X 轴和 Y 轴的主要和次要网格线
                    for axis in ['x_axis', 'y_axis']:
                        if hasattr(chart, axis):
                            axis_obj = getattr(chart, axis)
                            # 确保网格线被移除
                            axis_obj.majorGridlines = None
                            axis_obj.minorGridlines = None
            except Exception as e:
                logger.debug(f"XML 操作移除网格线失败（可忽略）: {e}")
            
            logger.info("图表已创建（已移除网格线）")
        except Exception as e:
            logger.error(f"创建图表失败: {str(e)}")
            import traceback
            traceback.print_exc()
            # 即使图表创建失败，也要保存数据
    
    def create_status_eval_report(
        self, 
        results: Dict[str, str], 
        evaluations: List['EvaluationItem'],
        output_path: str,
        assessment_content_map: Dict[str, str] = None
    ):
        """
        创建状态评估报表
        
        Args:
            results: 评估结果字典，{item_id: "是" 或 "否"}
            evaluations: 评估项配置列表（用于保持顺序和获取显示名称）
            output_path: 输出文件路径
            assessment_content_map: 评估内容描述映射，{item_id: "描述内容"}
        """
        if assessment_content_map is None:
            assessment_content_map = {}
        
        # 创建Excel工作簿
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "状态评估表"
        
        # 写入表头
        headers = ['评估项目', '评估内容', '评估结论']
        self.worksheet.append(headers)
        
        # 格式化表头：居中、加粗、背景色
        for col_idx in range(1, len(headers) + 1):
            cell = self.worksheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 遍历配置中的评估项（保持顺序）
        row_count = 0
        for eval_item in evaluations:
            # 跳过functional_result类型
            if eval_item.type == "functional_result":
                continue
            
            item_id = eval_item.item
            
            # 第一列：评估项目（使用assessmentName）
            assessment_name = eval_item.assessment_name if hasattr(eval_item, 'assessment_name') else item_id
            
            # 第二列：评估内容（从assessment_content_map获取）
            assessment_content = assessment_content_map.get(item_id, '')
            
            # 第三列：评估结论（从results字典获取）
            conclusion = results.get(item_id, "是")
            # 如果结果不在字典中，记录警告（虽然理论上不应该发生）
            if item_id not in results:
                logger.warning(f"评估项 {item_id} ({assessment_name}) 的结果不在results字典中，使用默认值'是'")
            # 确保结论值有效
            if conclusion not in ["是", "否"]:
                logger.warning(f"评估项 {item_id} ({assessment_name}) 的结论值无效: {conclusion}，使用默认值'是'")
                conclusion = "是"
            
            # 写入数据行
            row_values = [assessment_name, assessment_content, conclusion]
            self.worksheet.append(row_values)
            row_count += 1
            
            # 格式化数据行
            current_row = self.worksheet.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = self.worksheet.cell(row=current_row, column=col_idx)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # 评估结论列：如果是"否"，可以设置特殊颜色
                if col_idx == 3:  # 评估结论列（第三列）
                    if conclusion == "否":
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif conclusion == "是":
                        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        # 调整列宽
        self.worksheet.column_dimensions['A'].width = 30  # 评估项目
        self.worksheet.column_dimensions['B'].width = 50  # 评估内容
        self.worksheet.column_dimensions['C'].width = 15  # 评估结论
        
        logger.info(f"已写入 {row_count} 行状态评估数据")
        
        # 保存文件
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)
        logger.info(f"状态评估报表已保存: {output_path}")

