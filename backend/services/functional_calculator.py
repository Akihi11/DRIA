"""
功能计算汇总表 - 状态机实现 (库文件)
根据配置文件识别升速-降速循环并计算关键指标

V3.1 - 最终修复。
       - 修复了 V2.8 的 RAMPING_UP->RAMPING_DOWN 切换逻辑。
       - 现在通过检测 T_Baseline 通道的“峰值已过”（开始下降）
         来切换状态，确保能正确捕获 T1。
"""
import numpy as np
from collections import deque
from typing import List, Dict, Tuple, Any, Optional
from dataclasses import dataclass
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 日志记录器将由调用方（你的测试文件）配置
logger = logging.getLogger(__name__)


@dataclass
class FunctionalCalcConfig:
    """功能计算配置 (由调用方传入)"""
    time_base: Optional[Dict[str, Any]]
    startup_time: Optional[Dict[str, Any]]
    ignition_time: Optional[Dict[str, Any]]
    rundown_ng: Optional[Dict[str, Any]]
    rundown_np: Optional[Dict[str, Any]]


class SlidingWindow:
    """滑动窗口计算器"""
    
    def __init__(self, duration: float, statistic_type: str = "平均值"):
        self.duration = duration
        self.statistic_type = statistic_type
        self.window = deque()  # [(timestamp, value), ...]
    
    def update(self, timestamp: float, value: float):
        """更新窗口，移除过期数据，添加新数据"""
        # (V2.9 修复)
        cutoff_time_exclusive = timestamp - self.duration
        
        while self.window and self.window[0][0] <= cutoff_time_exclusive:
            self.window.popleft()
            
        self.window.append((timestamp, value))

    
    def calculate_statistic(self) -> Optional[float]:
        """计算窗口内的统计值"""
        if not self.window:
            return None
        values = [item[1] for item in self.window]
        array = np.array(values)
        stat_lower = self.statistic_type.lower() if self.statistic_type else 'average'
        
        if stat_lower in ['average', '平均值', 'mean', 'avg']:
            return float(np.mean(array))
        elif stat_lower in ['max', '最大值', 'maximum']:
            return float(np.max(array))
        elif stat_lower in ['min', '最小值', 'minimum']:
            return float(np.min(array))
        elif stat_lower in ['rms', '有效值', 'rootmeansquare']:
            return float(np.sqrt(np.mean(array ** 2)))
        else:
            logger.warning(f"未知的统计类型: {self.statistic_type}，使用平均值代替")
            return float(np.mean(array))
    
    def get_oldest_value(self) -> Optional[float]:
        if not self.window:
            return None
        return self.window[0][1]
    
    def get_oldest_time(self) -> Optional[float]:
        if not self.window:
            return None
        return self.window[0][0]


class FunctionalCalculator:
    """功能计算器 - 状态机实现"""
    
    def __init__(self, config: FunctionalCalcConfig):
        self.config = config
        self.state = "IDLE"
        self.startup_count = 0
        self.excel_row_index = 2
        self.current_cycle_data = {}
        
        self.windows = {}
        self.difference_windows = {}
        
        self._window_keys_to_channels = []
        self._diff_window_keys_to_channels = []

        # (V2.6 修复)
        self._startup_condition_was_true = False

        # --- (!!!) V3.1 修复：增加“峰值检测”逻辑 ---
        self._baseline_channel_key = None
        self._last_baseline_value = -np.inf # 必须初始化得非常小
        # --- (修复结束) ---

        if config.time_base:
            channel = config.time_base.get('channel')
            statistic = config.time_base.get('statistic', '平均值')
            duration = config.time_base.get('duration', 1.0)
            key = f"time_base_{channel}"
            self.windows[key] = SlidingWindow(duration, statistic)
            self._window_keys_to_channels.append((key, channel))
            self._baseline_channel_key = key # (V3.1) 记住T_Baseline的窗口key
        
        if config.startup_time:
            channel = config.startup_time.get('channel')
            statistic = config.startup_time.get('statistic', '平均值')
            duration = config.startup_time.get('duration', 1.0)
            key = f"startup_time_{channel}"
            self.windows[key] = SlidingWindow(duration, statistic)
            self._window_keys_to_channels.append((key, channel))
        
        if config.ignition_time:
            channel = config.ignition_time.get('channel')
            duration = config.ignition_time.get('duration', 10.0)
            key = f"ignition_time_{channel}"
            self.difference_windows[key] = SlidingWindow(duration, '瞬时值')
            self._diff_window_keys_to_channels.append((key, channel))
        
        if config.rundown_ng:
            channel = config.rundown_ng.get('channel')
            statistic = config.rundown_ng.get('statistic', '平均值')
            duration = config.rundown_ng.get('duration', 1.0)
            key = f"rundown_ng_{channel}"
            self.windows[key] = SlidingWindow(duration, statistic)
            self._window_keys_to_channels.append((key, channel))
        
        if config.rundown_np:
            channel = config.rundown_np.get('channel')
            statistic = config.rundown_np.get('statistic', '平均值')
            duration = config.rundown_np.get('duration', 1.0)
            key = f"rundown_np_{channel}"
            self.windows[key] = SlidingWindow(duration, statistic)
            self._window_keys_to_channels.append((key, channel))
        
        self.results = []
        logger.info(f"功能计算器初始化完成，状态机状态: {self.state}")
    
    def _update_all_windows(self, timestamp: float, data_point: Dict[str, float]):
        """在循环顶部统一更新所有窗口"""
        for key, channel in self._window_keys_to_channels:
            if channel in data_point:
                self.windows[key].update(timestamp, data_point[channel])
        
        for key, channel in self._diff_window_keys_to_channels:
            if channel in data_point:
                self.difference_windows[key].update(timestamp, data_point[channel])
                
    def evaluate_logic(self, value: float, logic: str, threshold: float) -> bool:
        """评估逻辑条件"""
        if logic == ">":
            return value > threshold
        elif logic == "<":
            return value < threshold
        elif logic == ">=":
            return value >= threshold
        elif logic == "<=":
            return value <= threshold
        else:
            raise ValueError(f"不支持的逻辑操作: {logic}")
    
    def is_startup_time_met(self) -> bool:
        """检查startup_time条件是否满足"""
        if not self.config.startup_time:
            return False
        config = self.config.startup_time
        channel = config.get('channel')
        key = f"startup_time_{channel}"
        if key not in self.windows:
            return False
        window = self.windows[key]
        stat_value = window.calculate_statistic()
        if stat_value is None:
            return False
        logic = config.get('logic', '>')
        threshold = config.get('threshold', 0.0)
        result = self.evaluate_logic(stat_value, logic, threshold)
        logger.debug(f"is_startup_time_met: ({key} Val: {stat_value:.2f}) {logic} {threshold}? -> {result}")
        return result
    
    def is_time_base_met(self) -> bool:
        """检查time_base条件是否满足"""
        if not self.config.time_base:
            return False
        config = self.config.time_base
        channel = config.get('channel')
        key = f"time_base_{channel}"
        if key not in self.windows:
            return False
        window = self.windows[key]
        stat_value = window.calculate_statistic()
        if stat_value is None:
            return False
        logic = config.get('logic', '>')
        threshold = config.get('threshold', 0.0)
        result = self.evaluate_logic(stat_value, logic, threshold)
        logger.debug(f"is_time_base_met: ({key} Val: {stat_value:.2f}) {logic} {threshold}? -> {result}")
        return result

    
    def is_ignition_time_met(self, data_point: Dict[str, float]) -> bool:
        """检查ignition_time条件是否满足（差值计算）"""
        if not self.config.ignition_time:
            return False
        config = self.config.ignition_time
        channel = config.get('channel')
        if channel not in data_point:
            return False
        key = f"ignition_time_{channel}"
        if key not in self.difference_windows:
            return False
        window = self.difference_windows[key]
        current_value = data_point[channel]
        oldest_value = window.get_oldest_value()
        if oldest_value is None:
            return False
        difference = current_value - oldest_value
        logic = config.get('logic', '>')
        threshold = config.get('threshold', 0.0)
        result = self.evaluate_logic(difference, logic, threshold)
        logger.debug(f"is_ignition_time_met: ({key} Diff: {difference:.2f}) {logic} {threshold}? -> {result}")
        return result

    
    def is_ng_rundown_T1_met(self) -> bool:
        """检查rundown_ng的T1条件（threshold1）是否满足"""
        if not self.config.rundown_ng:
            return False
        config = self.config.rundown_ng
        channel = config.get('channel')
        key = f"rundown_ng_{channel}"
        if key not in self.windows:
            return False
        window = self.windows[key]
        stat_value = window.calculate_statistic()
        if stat_value is None:
            return False
        threshold1 = config.get('threshold1', 0.0)
        result = stat_value < threshold1
        logger.debug(f"is_ng_rundown_T1_met: ({key} Val: {stat_value:.2f}) < {threshold1}? -> {result}")
        return result

    
    def is_ng_rundown_T2_met(self) -> bool:
        """检查rundown_ng的T2条件（threshold2）是否满足"""
        if not self.config.rundown_ng:
            return False
        config = self.config.rundown_ng
        channel = config.get('channel')
        key = f"rundown_ng_{channel}"
        if key not in self.windows:
            return False
        window = self.windows[key]
        stat_value = window.calculate_statistic()
        if stat_value is None:
            return False
        threshold2 = config.get('threshold2', 0.0)
        result = stat_value < threshold2
        logger.debug(f"is_ng_rundown_T2_met: ({key} Val: {stat_value:.2f}) < {threshold2}? -> {result}")
        return result

    
    def is_np_rundown_T1_met(self) -> bool:
        """检查rundown_np的T1条件（threshold1）是否满足"""
        if not self.config.rundown_np:
            return False
        config = self.config.rundown_np
        channel = config.get('channel')
        key = f"rundown_np_{channel}"
        if key not in self.windows:
            return False
        window = self.windows[key]
        stat_value = window.calculate_statistic()
        if stat_value is None:
            return False
        threshold1 = config.get('threshold1', 0.0)
        result = stat_value < threshold1
        logger.debug(f"is_np_rundown_T1_met: ({key} Val: {stat_value:.2f}) < {threshold1}? -> {result}")
        return result

    
    def is_np_rundown_T2_met(self) -> bool:
        """检查rundown_np的T2条件（threshold2）是否满足"""
        if not self.config.rundown_np:
            return False
        config = self.config.rundown_np
        channel = config.get('channel')
        key = f"rundown_np_{channel}"
        if key not in self.windows:
            return False
        window = self.windows[key]
        stat_value = window.calculate_statistic()
        if stat_value is None:
            return False
        threshold2 = config.get('threshold2', 0.0)
        result = stat_value < threshold2
        logger.debug(f"is_np_rundown_T2_met: ({key} Val: {stat_value:.2f}) < {threshold2}? -> {result}")
        return result

    
    def all_enabled_t2_events_found(self) -> bool:
        """检查所有启用的T2事件是否都已找到"""
        if self.config.rundown_ng:
            if 'T_Ng_T2' not in self.current_cycle_data:
                return False
        
        if self.config.rundown_np:
            if 'T_Np_T2' not in self.current_cycle_data:
                return False
        
        if not self.config.rundown_ng and not self.config.rundown_np:
            if self.current_cycle_data:
                return True
            return False
        
        return True
    
    def calculate_row(self):
        """计算并准备写入Excel的行数据"""
        self.startup_count += 1
        
        calc_time_base = self.current_cycle_data.get('T_Baseline')
        calc_start_time = None
        calc_ignition_time = None
        calc_ng_rundown = None
        calc_np_rundown = None
        
        if calc_time_base is not None and self.current_cycle_data.get('T_Start') is not None:
            calc_start_time = calc_time_base - self.current_cycle_data['T_Start']
        
        if calc_time_base is not None and self.current_cycle_data.get('T_Ignition') is not None:
            calc_ignition_time = self.current_cycle_data['T_Ignition'] - calc_time_base
        
        if self.current_cycle_data.get('T_Ng_T1') is not None and self.current_cycle_data.get('T_Ng_T2') is not None:
            calc_ng_rundown = self.current_cycle_data['T_Ng_T2'] - self.current_cycle_data['T_Ng_T1']
        
        if self.current_cycle_data.get('T_Np_T1') is not None and self.current_cycle_data.get('T_Np_T2') is not None:
            calc_np_rundown = self.current_cycle_data['T_Np_T2'] - self.current_cycle_data['T_Np_T1']
        
        new_row = {
            'startup_count': self.startup_count,
            'time_base': calc_time_base,
            'startup_time': calc_start_time,
            'ignition_time': calc_ignition_time,
            'ng_rundown': calc_ng_rundown,
            'np_rundown': calc_np_rundown
        }
        
        self.results.append(new_row)
        
        logger.info(f"计算完成循环 #{self.startup_count}: "
                    f"时间={calc_time_base}, 启动时间={calc_start_time}, "
                    f"点火时间={calc_ignition_time}, Ng余转={calc_ng_rundown}, "
                    f"Np余转={calc_np_rundown}")
        
        # (V2.5 修复)
        logger.debug("Resetting all sliding windows for the next cycle...")
        for window in self.windows.values():
            window.window.clear()
        for window in self.difference_windows.values():
            window.window.clear()
        
        # (V3.1 修复)
        self._last_baseline_value = -np.inf

        self.excel_row_index += 1
        self.current_cycle_data = {}
        self.state = "IDLE"
    
    def process_data_stream(self, data_stream: List[Tuple[float, Dict[str, float]]]):
        """处理数据流，执行状态机逻辑"""
        logger.info(f"开始处理数据流，总点数: {len(data_stream)}")
        
        for timestamp, data_point in data_stream:
            
            # 1. 统一更新所有窗口
            self._update_all_windows(timestamp, data_point)
            
            is_startup_met = self.is_startup_time_met()
            
            # (V2.8 修复)
            is_baseline_met = self.is_time_base_met()
            
            # 2. 状态机逻辑
            if self.state == "IDLE":
                # (V2.6 修复)
                if is_startup_met and not self._startup_condition_was_true:
                    self.current_cycle_data['T_Start'] = timestamp
                    self.state = "RAMPING_UP"
                    logger.info(f"[IDLE->RAMPING_UP] T_Start={timestamp:.3f}s")
            
            elif self.state == "RAMPING_UP":
                
                # --- (!!!) 核心修复 V3.1 (开始) ---
                
                # 1. 捕获升速事件
                if 'T_Start' not in self.current_cycle_data and is_startup_met:
                    self.current_cycle_data['T_Start'] = timestamp
                
                if 'T_Baseline' not in self.current_cycle_data and is_baseline_met:
                    self.current_cycle_data['T_Baseline'] = timestamp
                    logger.info(f"[RAMPING_UP] T_Baseline={timestamp:.3f}s")
                
                if 'T_Ignition' not in self.current_cycle_data and self.is_ignition_time_met(data_point):
                    self.current_cycle_data['T_Ignition'] = timestamp
                    logger.info(f"[RAMPING_UP] T_Ignition={timestamp:.3f}s")
                
                # 2. 检查是否切换到 RAMPING_DOWN
                #    必须先确保 T_Baseline 至少被抓到过
                if 'T_Baseline' in self.current_cycle_data:
                    
                    # (V3.1) 获取 T_Baseline 窗口的 *当前* 值
                    current_baseline_value = self.windows[self._baseline_channel_key].calculate_statistic()
                    if current_baseline_value is None: current_baseline_value = -np.inf
                    
                    # 检查是否开始降速 (当前值 < 上一个值)
                    is_decreasing = (current_baseline_value < self._last_baseline_value)
                    
                    if is_decreasing:
                        self.state = "RAMPING_DOWN"
                        logger.info(f"[RAMPING_UP->RAMPING_DOWN] Peak detected. T={timestamp:.3f}s")
                    
                    # 更新 "上一个值"
                    self._last_baseline_value = current_baseline_value

                # --- (!!!) 核心修复 V3.1 (结束) ---
            
            elif self.state == "RAMPING_DOWN":
                
                # (V2.7 修复)
                
                # 捕获Ng T1
                if 'T_Ng_T1' not in self.current_cycle_data and self.is_ng_rundown_T1_met():
                    self.current_cycle_data['T_Ng_T1'] = timestamp
                    logger.info(f"[RAMPING_DOWN] T_Ng_T1={timestamp:.3f}s")

                # 捕获Ng T2
                if 'T_Ng_T2' not in self.current_cycle_data and self.is_ng_rundown_T2_met():
                    self.current_cycle_data['T_Ng_T2'] = timestamp
                    logger.info(f"[RAMPING_DOWN] T_Ng_T2={timestamp:.3f}s")
                
                # 捕获Np T1
                if 'T_Np_T1' not in self.current_cycle_data and self.is_np_rundown_T1_met():
                    self.current_cycle_data['T_Np_T1'] = timestamp
                    logger.info(f"[RAMPING_DOWN] T_Np_T1={timestamp:.3f}s")
                
                # 捕获Np T2
                if 'T_Np_T2' not in self.current_cycle_data and self.is_np_rundown_T2_met():
                    self.current_cycle_data['T_Np_T2'] = timestamp
                    logger.info(f"[RAMPING_DOWN] T_Np_T2={timestamp:.3f}s")
                
                # 补齐升速阶段未捕获的事件
                if 'T_Baseline' not in self.current_cycle_data and is_baseline_met:
                    self.current_cycle_data['T_Baseline'] = timestamp
                
                if 'T_Ignition' not in self.current_cycle_data and self.is_ignition_time_met(data_point):
                    self.current_cycle_data['T_Ignition'] = timestamp
                
                # 检查循环是否结束
                if self.all_enabled_t2_events_found():
                    self.state = "CALCULATE_ROW"
            
            elif self.state == "CALCULATE_ROW":
                self.calculate_row()
                # 状态已在 calculate_row() 中切换到 IDLE
            
            # (V2.6 修复) 实时更新 "记忆"
            self._startup_condition_was_true = is_startup_met

        
        # 数据流结束时的处理
        if self.state in ["RAMPING_UP", "RAMPING_DOWN"] and self.current_cycle_data:
            if self.all_enabled_t2_events_found():
                logger.warning("数据流结束，强制计算最后一个循环")
                self.state = "CALCULATE_ROW"
                self.calculate_row()
            else:
                 logger.warning("数据流结束，但最后一个循环未完成（T2未找到），已丢弃")
        
        logger.info(f"数据处理完成，共识别 {len(self.results)} 个循环")
    
    def export_to_excel(self, output_path: str):
        """导出结果到Excel文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "功能计算汇总表"
        
        headers = ["启动次数", "时间", "启动时间", "点火时间", "Ng余转时间", "Np余转时间"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_data in self.results:
            
            # (V.2.3 版修改)
            def format_value(value):
                if value is None:
                    return ""
                else:
                    return value
            
            row = [
                row_data['startup_count'],
                format_value(row_data['time_base']),
                format_value(row_data['startup_time']),
                format_value(row_data['ignition_time']),
                format_value(row_data['ng_rundown']),
                format_value(row_data['np_rundown'])
            ]
            ws.append(row)
        
        column_widths = [12, 15, 15, 15, 15, 15]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        wb.save(output_path)
        logger.info(f"Excel文件已保存到: {output_path}")
        
        return output_path