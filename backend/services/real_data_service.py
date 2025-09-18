"""
Real implementations of data service interfaces for Phase 2
Implements actual CSV/Excel data reading and Excel report writing
"""
from typing import List, Dict, Any, Optional
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime
import logging

import sys
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent))

from interfaces.data_interfaces import DataReader, ReportWriter
from models.data_models import ChannelData, DataPoint, ReportData, FileMetadata

logger = logging.getLogger(__name__)


class RealDataReader(DataReader):
    """Real implementation of DataReader using pandas"""
    
    def __init__(self):
        self.supported_formats = {'.csv', '.xlsx', '.xls'}
    
    def read(self, file_path: str, channel_names: List[str]) -> List[ChannelData]:
        """
        Read actual data from CSV/Excel files
        
        Args:
            file_path: Path to the data file
            channel_names: List of channel names to read
            
        Returns:
            List[ChannelData]: List of channel data objects
        """
        try:
            # Read the file based on extension
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.csv':
                # Try different encodings for CSV files
                try:
                    df = pd.read_csv(file_path, encoding='utf-8')
                except UnicodeDecodeError:
                    try:
                        df = pd.read_csv(file_path, encoding='gbk')
                    except UnicodeDecodeError:
                        df = pd.read_csv(file_path, encoding='latin-1')
            elif file_ext in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            logger.info(f"Successfully read file {file_path} with shape {df.shape}")
            
            # Get time column (assume first column is time)
            time_column = df.columns[0]
            time_values = df[time_column].values
            
            # Calculate sample rate
            if len(time_values) > 1:
                sample_rate = 1.0 / (time_values[1] - time_values[0])
            else:
                sample_rate = 1.0
            
            channels = []
            
            # Map channel names to actual column names in the file
            column_mapping = self._map_channel_names(df.columns.tolist(), channel_names)
            
            for channel_name in channel_names:
                if channel_name in column_mapping:
                    actual_column = column_mapping[channel_name]
                    values = df[actual_column].values
                    
                    # Create data points
                    data_points = []
                    for i, (time_val, value) in enumerate(zip(time_values, values)):
                        data_points.append(DataPoint(
                            timestamp=float(time_val),
                            value=float(value) if pd.notna(value) else 0.0
                        ))
                    
                    # Create channel data
                    channel_data = ChannelData(
                        channel_name=channel_name,
                        unit=self._get_channel_unit(channel_name),
                        data_points=data_points,
                        sample_rate=sample_rate
                    )
                    
                    channels.append(channel_data)
                    logger.info(f"Loaded channel {channel_name} with {len(data_points)} data points")
                else:
                    logger.warning(f"Channel {channel_name} not found in file")
            
            return channels
            
        except Exception as e:
            logger.error(f"Error reading file {file_path}: {e}")
            raise
    
    def get_available_channels(self, file_path: str) -> List[str]:
        """Get all available channels in the file"""
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.csv':
                # Try different encodings for CSV files
                try:
                    df = pd.read_csv(file_path, nrows=0, encoding='utf-8')  # Read only headers
                except UnicodeDecodeError:
                    try:
                        df = pd.read_csv(file_path, nrows=0, encoding='gbk')
                    except UnicodeDecodeError:
                        df = pd.read_csv(file_path, nrows=0, encoding='latin-1')
            elif file_ext in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path, nrows=0)  # Read only headers
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            # Exclude time column (assume first column is time)
            channels = df.columns.tolist()[1:]
            return channels
            
        except Exception as e:
            logger.error(f"Error reading file headers {file_path}: {e}")
            raise
    
    def get_file_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract comprehensive file metadata"""
        try:
            file_path_obj = Path(file_path)
            file_ext = file_path_obj.suffix.lower()
            
            if file_ext == '.csv':
                # Try different encodings for CSV files
                try:
                    df = pd.read_csv(file_path, encoding='utf-8')
                except UnicodeDecodeError:
                    try:
                        df = pd.read_csv(file_path, encoding='gbk')
                    except UnicodeDecodeError:
                        df = pd.read_csv(file_path, encoding='latin-1')
            elif file_ext in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            # Calculate duration and sample rate
            time_column = df.columns[0]
            time_values = df[time_column].values
            
            duration = float(time_values[-1] - time_values[0]) if len(time_values) > 1 else 0.0
            sample_rate = len(time_values) / duration if duration > 0 else 0.0
            
            return {
                "filename": file_path_obj.name,
                "file_size": file_path_obj.stat().st_size,
                "total_samples": len(df),
                "duration_seconds": duration,
                "sample_rate": sample_rate,
                "channels_count": len(df.columns) - 1,  # Exclude time column
                "channels": df.columns.tolist()[1:],
                "file_format": file_ext,
                "created_time": datetime.fromtimestamp(file_path_obj.stat().st_mtime).isoformat(),
                "time_range": {
                    "start": float(time_values[0]),
                    "end": float(time_values[-1])
                }
            }
            
        except Exception as e:
            logger.error(f"Error extracting metadata from {file_path}: {e}")
            raise
    
    def _map_channel_names(self, file_columns: List[str], requested_channels: List[str]) -> Dict[str, str]:
        """
        Map requested channel names to actual column names in the file
        Handles different naming conventions
        """
        mapping = {}
        
        # Create a mapping of normalized names to actual names
        file_columns_normalized = {col.lower().replace(' ', '').replace('_', '').replace('-', ''): col 
                                 for col in file_columns}
        
        for requested in requested_channels:
            # Try exact match first
            if requested in file_columns:
                mapping[requested] = requested
                continue
            
            # Try normalized matching
            requested_normalized = requested.lower().replace(' ', '').replace('_', '').replace('-', '')
            
            # Special mappings for common variations
            special_mappings = {
                'ng(rpm)': ['ng', 'n1', 'n1(rpm)', 'ng(rpm)', 'ng转速'],
                'np(rpm)': ['np', 'n2', 'n2(rpm)', 'np(rpm)', 'np转速'],
                'temperature(°c)': ['temperature', 'temp', '温度', '排气温度', 'egt'],
                'pressure(kpa)': ['pressure', 'press', '压力', '滑油压力', 'oil_pressure']
            }
            
            found = False
            for pattern, variations in special_mappings.items():
                if requested_normalized in pattern or any(v in requested_normalized for v in variations):
                    # Find matching column
                    for file_col_norm, file_col_actual in file_columns_normalized.items():
                        if any(v in file_col_norm for v in variations):
                            mapping[requested] = file_col_actual
                            found = True
                            break
                    if found:
                        break
            
            if not found:
                # Try partial matching
                for file_col_norm, file_col_actual in file_columns_normalized.items():
                    if requested_normalized in file_col_norm or file_col_norm in requested_normalized:
                        mapping[requested] = file_col_actual
                        break
        
        return mapping
    
    def _get_channel_unit(self, channel_name: str) -> str:
        """Get unit for channel based on name"""
        unit_mappings = {
            'ng': 'rpm',
            'np': 'rpm',
            'temperature': '°C',
            'pressure': 'kPa',
            'voltage': 'V',
            'current': 'A',
            'fuel': 'kg/h',
            'vibration': 'mm/s'
        }
        
        channel_lower = channel_name.lower()
        for key, unit in unit_mappings.items():
            if key in channel_lower:
                return unit
        
        # Extract unit from channel name if present (e.g., "Temperature(°C)")
        if '(' in channel_name and ')' in channel_name:
            return channel_name.split('(')[1].split(')')[0]
        
        return ""


class RealReportWriter(ReportWriter):
    """Real implementation of ReportWriter using openpyxl"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def write(self, output_path: str, report_data: ReportData) -> bool:
        """Write report data to Excel file"""
        try:
            return self.create_excel_report(output_path, report_data)
        except Exception as e:
            self.logger.error(f"Error writing report to {output_path}: {e}")
            return False
    
    def create_excel_report(self, output_path: str, report_data: ReportData) -> bool:
        """Create Excel report with multiple sheets"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.chart import LineChart, Reference
            
            # Create workbook and remove default sheet
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Define styles
            header_font = Font(bold=True, size=12)
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Create summary sheet
            summary_ws = wb.create_sheet("报表摘要")
            self._create_summary_sheet(summary_ws, report_data, title_font, header_font, border)
            
            # Create stable state sheet if data exists
            if report_data.stable_state_result:
                stable_ws = wb.create_sheet("稳定状态参数汇总")
                self._create_stable_state_sheet(stable_ws, report_data.stable_state_result, 
                                              title_font, header_font, border)
            
            # Create functional calc sheet if data exists
            if report_data.functional_calc_result:
                func_ws = wb.create_sheet("功能计算汇总")
                self._create_functional_calc_sheet(func_ws, report_data.functional_calc_result,
                                                 title_font, header_font, border)
            
            # Create status evaluation sheet if data exists
            if report_data.status_eval_result:
                status_ws = wb.create_sheet("状态评估")
                self._create_status_eval_sheet(status_ws, report_data.status_eval_result,
                                             title_font, header_font, border)
            
            # Save workbook
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            
            self.logger.info(f"Successfully created Excel report: {output_path}")
            return True
            
        except ImportError:
            self.logger.error("openpyxl not installed. Please install: pip install openpyxl")
            return False
        except Exception as e:
            self.logger.error(f"Error creating Excel report: {e}")
            return False
    
    def _create_summary_sheet(self, ws, report_data, title_font, header_font, border):
        """Create summary sheet with basic report information"""
        ws['A1'] = "AI报表生成系统 - 分析报告"
        ws['A1'].font = title_font
        
        ws['A3'] = "报表基本信息"
        ws['A3'].font = header_font
        
        info_data = [
            ["报表ID", report_data.report_id],
            ["源文件ID", report_data.source_file_id],
            ["生成时间", report_data.generation_time.strftime("%Y-%m-%d %H:%M:%S")],
            ["包含内容", ""],
        ]
        
        row = 4
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = header_font
            row += 1
        
        # Add content summary
        if report_data.stable_state_result:
            ws[f'B{row-1}'] = ws[f'B{row-1}'].value + "稳定状态分析; "
        if report_data.functional_calc_result:
            ws[f'B{row-1}'] = ws[f'B{row-1}'].value + "功能计算分析; "
        if report_data.status_eval_result:
            ws[f'B{row-1}'] = ws[f'B{row-1}'].value + "状态评估分析; "
    
    def _create_stable_state_sheet(self, ws, stable_result, title_font, header_font, border):
        """Create stable state analysis sheet"""
        ws['A1'] = "稳定状态参数汇总表"
        ws['A1'].font = title_font
        
        # Summary statistics
        ws['A3'] = "分析摘要"
        ws['A3'].font = header_font
        
        ws['A4'] = "总稳定时间(秒)"
        ws['B4'] = stable_result.total_stable_time
        ws['A5'] = "稳定时段数量"
        ws['B5'] = len(stable_result.stable_periods)
        
        # Stable periods table
        ws['A7'] = "稳定时段详情"
        ws['A7'].font = header_font
        
        headers = ["时段序号", "开始时间(s)", "结束时间(s)", "持续时间(s)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=8, column=col, value=header).font = header_font
        
        for i, period in enumerate(stable_result.stable_periods, 1):
            ws.cell(row=8+i, column=1, value=i)
            ws.cell(row=8+i, column=2, value=period.get('start_time', 0))
            ws.cell(row=8+i, column=3, value=period.get('end_time', 0))
            ws.cell(row=8+i, column=4, value=period.get('duration', 0))
    
    def _create_functional_calc_sheet(self, ws, func_result, title_font, header_font, border):
        """Create functional calculation sheet"""
        ws['A1'] = "功能计算汇总表"
        ws['A1'].font = title_font
        
        ws['A3'] = "计算结果"
        ws['A3'].font = header_font
        
        results = [
            ["时间基准(秒)", func_result.time_base],
            ["启动时间(秒)", func_result.startup_time],
            ["点火时间(秒)", func_result.ignition_time],
            ["Ng余转时间(秒)", func_result.rundown_ng]
        ]
        
        # Headers
        ws['A4'] = "指标名称"
        ws['B4'] = "计算结果"
        ws['A4'].font = header_font
        ws['B4'].font = header_font
        
        row = 5
        for metric, value in results:
            if value is not None:
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = f"{value:.2f}"
                row += 1
    
    def _create_status_eval_sheet(self, ws, status_result, title_font, header_font, border):
        """Create status evaluation sheet"""
        ws['A1'] = "状态评估表"
        ws['A1'].font = title_font
        
        ws['A3'] = f"总体状态: {status_result.overall_status}"
        ws['A3'].font = header_font
        
        # Evaluation results table
        ws['A5'] = "详细评估结果"
        ws['A5'].font = header_font
        
        headers = ["评估项", "评估结果", "状态"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=6, column=col, value=header).font = header_font
        
        row = 7
        for eval_item in status_result.evaluations:
            ws.cell(row=row, column=1, value=eval_item.get('item', ''))
            ws.cell(row=row, column=2, value=eval_item.get('result', ''))
            ws.cell(row=row, column=3, value=eval_item.get('status', ''))
            row += 1
        
        # Warnings section
        if status_result.warnings:
            ws[f'A{row+1}'] = "警告信息"
            ws[f'A{row+1}'].font = header_font
            
            for i, warning in enumerate(status_result.warnings):
                ws[f'A{row+2+i}'] = f"• {warning}"
    
    def add_charts_to_report(self, report_path: str, chart_configs: List[Dict[str, Any]]) -> bool:
        """Add charts to existing Excel report"""
        try:
            import openpyxl
            from openpyxl.chart import LineChart, Reference
            
            wb = openpyxl.load_workbook(report_path)
            
            for config in chart_configs:
                sheet_name = config.get('sheet', '报表摘要')
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Create chart based on config
                    chart = LineChart()
                    chart.title = config.get('title', 'Chart')
                    chart.y_axis.title = config.get('y_label', 'Values')
                    chart.x_axis.title = config.get('x_label', 'Time')
                    
                    # Add data series (simplified implementation)
                    data_range = config.get('data_range', 'A1:B10')
                    data = Reference(ws, range_string=data_range)
                    chart.add_data(data, titles_from_data=True)
                    
                    # Add chart to worksheet
                    position = config.get('position', 'D2')
                    ws.add_chart(chart, position)
            
            wb.save(report_path)
            return True
            
        except Exception as e:
            self.logger.error(f"Error adding charts to report: {e}")
            return False
